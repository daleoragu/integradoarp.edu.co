# notas/reportes/excel_generator.py
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from openpyxl.utils import get_column_letter

from .base_generator import BaseReportGenerator
# --- CORRECCIÓN: Se importa la función específica para la plantilla ---
from .utils import get_meses_for_periodo, get_asistencia_data_for_template

class AsistenciaExcelGenerator(BaseReportGenerator):
    """
    Genera una PLANTILLA de asistencia en Excel con encabezado dinámico.
    """
    def generate_report(self, asignacion, periodo, mes_seleccionado: str):
        wb = Workbook()
        wb.remove(wb.active)
        
        meses_a_procesar = []
        if mes_seleccionado and mes_seleccionado.lower() != 'todos':
            try:
                mes_num = int(mes_seleccionado)
                nombre_mes = datetime.date(periodo.fecha_inicio.year, mes_num, 1).strftime('%B %Y').capitalize()
                meses_a_procesar.append((mes_num, nombre_mes))
            except (ValueError, TypeError):
                pass
        else:
            meses_a_procesar = get_meses_for_periodo(periodo)

        if not meses_a_procesar:
            ws = wb.create_sheet(title="Sin Datos")
            self._add_excel_header(ws)
            ws['A7'] = "No hay meses configurados para este periodo."
            return wb

        for mes_num, nombre_completo in meses_a_procesar:
            # --- CORRECCIÓN: Llama a la nueva función para obtener los días hábiles ---
            estudiantes, fechas, resumen_vacio = get_asistencia_data_for_template(asignacion, periodo, mes_num)
            
            nombre_hoja = nombre_completo.split(' ')[0][:30]
            ws = wb.create_sheet(title=nombre_hoja)
            
            # El resumen va vacío, por lo que las celdas quedarán en blanco
            self._write_sheet(ws, asignacion, estudiantes, fechas, resumen_vacio, nombre_completo)
                
        return wb

    def _write_sheet(self, ws, asignacion, estudiantes, fechas_del_mes, resumen_asistencia, nombre_mes_año):
        """
        Escribe la hoja de cálculo. Como resumen_asistencia está vacío,
        las celdas de asistencia quedarán en blanco.
        """
        ws.protection.sheet = True
        self._add_excel_header(ws)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center')

        ws.merge_cells('A7:P7')
        ws['A7'] = "PLANTILLA DE ASISTENCIA MENSUAL"
        ws['A7'].font = Font(bold=True, size=14); ws['A7'].alignment = center_align
        
        ws.merge_cells('A8:P8')
        docente_nombre = asignacion.docente.user.get_full_name() if asignacion.docente else "No asignado"
        info_reporte_text = f"CURSO: {asignacion.curso.nombre}   |   ASIGNATURA: {asignacion.materia.nombre}   |   DOCENTE: {docente_nombre}"
        ws['A8'].value = info_reporte_text
        ws['A8'].font = Font(bold=True, size=11); ws['A8'].alignment = center_align
        
        for col_idx, fecha in enumerate(fechas_del_mes, start=3):
            ws.cell(row=9, column=col_idx, value=f"{asignacion.id}|{fecha.strftime('%Y-%m-%d')}")
        ws.row_dimensions[9].hidden = True

        ws.merge_cells('A10:B11')
        ws['A10'].value = "Estudiante"
        ws['A10'].font = header_font; ws['A10'].fill = header_fill; ws['A10'].alignment = left_align
        
        if fechas_del_mes:
            start_col, end_col = get_column_letter(3), get_column_letter(2 + len(fechas_del_mes))
            ws.merge_cells(f'{start_col}10:{end_col}10')
            cell_mes = ws.cell(row=10, column=3)
            cell_mes.value = nombre_mes_año.upper()
            cell_mes.font = header_font; cell_mes.fill = header_fill; cell_mes.alignment = center_align
            for col_idx, fecha in enumerate(fechas_del_mes, start=3):
                cell_dia = ws.cell(row=11, column=col_idx, value=fecha.day)
                cell_dia.font = header_font; cell_dia.fill = header_fill; cell_dia.alignment = center_align
        
        unlocked_protection = Protection(locked=False)
        start_data_row = 12 
        
        for row_offset, estudiante in enumerate(estudiantes):
            current_row_num = start_data_row + row_offset
            nombre_formateado = f"{estudiante.user.last_name.upper()} {estudiante.user.first_name.upper()}"
            ws.cell(row=current_row_num, column=1, value=nombre_formateado).alignment = left_align
            ws.cell(row=current_row_num, column=2, value=estudiante.id)
            
            asistencias_estudiante = resumen_asistencia.get(estudiante.id, {})
            for col_offset, fecha in enumerate(fechas_del_mes, start=3):
                # Esta línea ahora resultará en un string vacío, dejando la celda en blanco
                estado_asistencia = asistencias_estudiante.get(fecha, '')
                cell_asistencia = ws.cell(row=current_row_num, column=col_offset, value=estado_asistencia)
                cell_asistencia.alignment = center_align
                cell_asistencia.protection = unlocked_protection
                
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].hidden = True
        if fechas_del_mes:
            for i in range(3, len(fechas_del_mes) + 3):
                ws.column_dimensions[get_column_letter(i)].width = 5
