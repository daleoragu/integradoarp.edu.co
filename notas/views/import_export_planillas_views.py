# notas/views/import_export_planillas_views.py

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
# Se añade HttpResponseNotFound para manejar errores
from django.http import HttpResponse, HttpResponseNotFound
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.conf import settings
import os

from ..models import AsignacionDocente, PeriodoAcademico, Estudiante, Docente

@login_required
def exportar_planillas_docente(request, docente_id, periodo_id):
    """
    Genera y descarga un único archivo Excel con una hoja por cada asignación,
    asegurando que los datos correspondan al colegio actual.
    """
    # CORRECCIÓN: Verificar que se ha identificado un colegio.
    if not request.colegio:
        return HttpResponseNotFound("<h1>Colegio no configurado</h1>")

    # CORRECCIÓN: Filtrar docente y periodo por el colegio actual para seguridad.
    docente = get_object_or_404(Docente, id=docente_id, colegio=request.colegio)
    periodo = get_object_or_404(PeriodoAcademico, id=periodo_id, colegio=request.colegio)
    
    # CORRECCIÓN: Filtrar las asignaciones también por el colegio actual para doble seguridad.
    asignaciones = AsignacionDocente.objects.filter(
        docente=docente, 
        colegio=request.colegio
    ).order_by('curso__nombre', 'materia__nombre')

    if not asignaciones.exists():
        return HttpResponse("Este docente no tiene asignaturas asignadas en este colegio.", status=404)

    workbook = openpyxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # MEJORA A FUTURO: Los logos deberían cargarse desde el objeto `request.colegio`
    # en lugar de una ruta estática para permitir personalización.
    # Por ahora, se mantiene la lógica original.
    try:
        # Usamos el logo del modelo Colegio si existe, si no, el de por defecto.
        if request.colegio.logo:
            logo_colegio_path = request.colegio.logo.path
        else:
            logo_colegio_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', 'logo_colegio.png')
        
        logo_gob_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', 'Logo_govtolima.png')
        logo_colegio_img = OpenpyxlImage(logo_colegio_path)
        logo_gob_img = OpenpyxlImage(logo_gob_path)
        logo_colegio_img.height = 65
        logo_colegio_img.width = 65
        logo_gob_img.height = 65
        logo_gob_img.width = 65
        logos_cargados = True
    except (FileNotFoundError, IndexError, AttributeError):
        logos_cargados = False
        print("ADVERTENCIA: No se encontraron los archivos de logo. El Excel se generará sin ellos.")

    for asignacion in asignaciones:
        sheet = workbook.create_sheet(title=f"{asignacion.curso.nombre}-{asignacion.materia.abreviatura}")
        
        # CORRECCIÓN: Filtrar estudiantes por el colegio actual.
        estudiantes = Estudiante.objects.filter(
            curso=asignacion.curso, 
            colegio=request.colegio, 
            is_active=True
        ).order_by('user__last_name', 'user__first_name')

        # --- ENCABEZADO INSTITUCIONAL ---
        font_titulo = Font(name='Arial', size=11, bold=True)
        font_info_label = Font(name='Arial', size=10, bold=True)
        font_info_value = Font(name='Arial', size=10)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        
        sheet.merge_cells('B1:C4')
        cell_titulo = sheet['B1']
        # Usamos el nombre del colegio del objeto `request.colegio`
        cell_titulo.value = (f"{request.colegio.nombre.upper()}\n"
                             f"Nit. [NIT del Colegio]\n" # Estos datos podrían ser campos en el modelo Colegio
                             f"DANE [DANE del Colegio]\n"
                             f"[Ciudad]")
        cell_titulo.font = font_titulo
        cell_titulo.alignment = center_align
        
        if logos_cargados:
            sheet.add_image(logo_gob_img, 'A1')
            sheet.add_image(logo_colegio_img, 'D1')
        
        info_row_start = 6
        info_data = {
            "Docente:": asignacion.docente.user.get_full_name(), "Grado:": asignacion.curso.nombre,
            "Año:": periodo.ano_lectivo, "Periodo:": periodo.get_nombre_display(),
            "Materia:": asignacion.materia.nombre,
        }
        
        row_idx = info_row_start
        for label, value in info_data.items():
            sheet.cell(row=row_idx, column=1, value=label).font = font_info_label
            sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
            sheet.cell(row=row_idx, column=2, value=value).font = font_info_value
            row_idx += 1
        
        # --- ESTRUCTURA DE TABLA DE NOTAS (sin cambios en la lógica de celdas) ---
        header_row = row_idx + 1
        sub_header_row = header_row + 1
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6A0000", end_color="6A0000", fill_type="solid")
        
        sheet.cell(row=header_row, column=1, value="ID Estudiante").font = header_font; sheet.cell(row=header_row, column=1).fill = header_fill
        sheet.cell(row=header_row, column=2, value="Apellidos y Nombres").font = header_font; sheet.cell(row=header_row, column=2).fill = header_fill
        sheet.merge_cells(start_row=header_row, start_column=1, end_row=sub_header_row, end_column=1)
        sheet.merge_cells(start_row=header_row, start_column=2, end_row=sub_header_row, end_column=2)
        
        current_col = 3
        num_notas_fijas = 5

        for dimension in ["SER", "SABER", "HACER"]:
            start_col = current_col
            end_col = start_col + num_notas_fijas - 1
            sheet.merge_cells(start_row=header_row, start_column=start_col, end_row=header_row, end_column=end_col)
            cell = sheet.cell(row=header_row, column=start_col, value=dimension)
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align

            for i in range(num_notas_fijas):
                sheet.cell(row=sub_header_row, column=current_col + i, value=f"n{i+1}").alignment = center_align
            current_col += num_notas_fijas
                
        sheet.column_dimensions['B'].width = 40

        for row_num, estudiante in enumerate(estudiantes, sub_header_row + 1):
            nombre_completo = f"{estudiante.user.last_name}, {estudiante.user.first_name}".strip()
            sheet.cell(row=row_num, column=1, value=estudiante.id)
            sheet.cell(row=row_num, column=2, value=nombre_completo)

        sheet.protection.sheet = True
        for row in sheet.iter_rows(min_row=sub_header_row + 1, max_row=sheet.max_row, min_col=3):
            for cell in row:
                cell.protection = openpyxl.styles.Protection(locked=False)

    nombre_archivo = f"Planilla de Notas - {periodo} - {docente.user.get_full_name()}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre_archivo}"'}
    )
    workbook.save(response)
    return response
