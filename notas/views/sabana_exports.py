# notas/views/sabana_exports.py

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import requests
from io import BytesIO
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone

def generar_excel_sabana(curso, periodo, sabana_data, areas_con_materias, desempenos_headers, resumen_final_celdas, colegio, is_final_report, **kwargs):
    """
    Genera un archivo Excel con la Sábana de Notas, con un diseño profesional
    y mostrando las notas originales junto a las de recuperación.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sabana {curso.nombre}"

    # --- Definición de Estilos ---
    title_font = Font(name='Calibri', bold=True, size=16)
    subtitle_font = Font(name='Calibri', bold=True, size=12)
    bold_font = Font(name='Calibri', bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    area_header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    acumulado_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # --- Encabezado del Reporte (similar al PDF) ---
    # Se calcula el número total de columnas para poder centrar el encabezado
    total_cols = 3 + sum(len(a.materias_del_curso_ordenadas) + 1 for a in areas_con_materias) + 1 + len(desempenos_headers) + 1
    
    # Insertar logo si existe
    if colegio.logo:
        try:
            response = requests.get(colegio.logo.url, stream=True)
            response.raise_for_status()
            img = Image(BytesIO(response.content))
            img.height = 60
            img.width = 60
            ws.add_image(img, 'A1')
        except requests.exceptions.RequestException:
            pass # Si falla la descarga del logo, simplemente no se añade.
    
    # Información del colegio
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols -1)
    cell = ws.cell(row=1, column=2, value=colegio.nombre.upper())
    cell.font = title_font
    cell.alignment = center_align

    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=total_cols -1)
    info_text = f"REGISTRO DANE: {colegio.dane or 'N/A'} | RESOLUCIÓN {colegio.resolucion or 'N/A'}"
    cell = ws.cell(row=2, column=2, value=info_text)
    cell.alignment = center_align

    # Título de la sábana
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=total_cols)
    title_text = f"SÁBANA DE NOTAS FINAL - {periodo.ano_lectivo}" if is_final_report else f"SÁBANA DE NOTAS ACUMULATIVA AL {periodo.get_nombre_display().upper()} - {periodo.ano_lectivo}"
    cell = ws.cell(row=4, column=1, value=title_text)
    cell.font = subtitle_font
    cell.alignment = center_align

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=total_cols)
    director = curso.director_grado.user.get_full_name() if curso.director_grado else "No asignado"
    cell = ws.cell(row=5, column=1, value=f"CURSO: {curso.nombre} | DIRECTOR: {director}")
    cell.alignment = center_align

    # --- Encabezados de la Tabla ---
    row_h1, row_h2 = 7, 8
    ws.cell(row=row_h1, column=1, value="N°")
    ws.cell(row=row_h1, column=2, value="Apellidos y Nombres")
    ws.cell(row=row_h1, column=3, value="P.")
    ws.merge_cells(start_row=row_h1, start_column=1, end_row=row_h2, end_column=1)
    ws.merge_cells(start_row=row_h1, start_column=2, end_row=row_h2, end_column=2)
    ws.merge_cells(start_row=row_h1, start_column=3, end_row=row_h2, end_column=3)

    col_idx = 4
    for area in areas_con_materias:
        num_materias = len(area.materias_del_curso_ordenadas)
        ws.cell(row=row_h1, column=col_idx, value=area.nombre).fill = area_header_fill
        if num_materias > 0:
            ws.merge_cells(start_row=row_h1, start_column=col_idx, end_row=row_h1, end_column=col_idx + num_materias)
        
        for materia in area.materias_del_curso_ordenadas:
            ws.cell(row=row_h2, column=col_idx, value=materia.abreviatura or materia.nombre)
            col_idx += 1
        ws.cell(row=row_h2, column=col_idx, value="DEF. ÁREA").fill = area_header_fill
        col_idx += 1
    
    prom_col = col_idx
    ws.cell(row=row_h1, column=prom_col, value="Prom.")
    ws.merge_cells(start_row=row_h1, start_column=prom_col, end_row=row_h2, end_column=prom_col)
    col_idx += 1

    for d in desempenos_headers:
        ws.cell(row=row_h1, column=col_idx, value=d[:3] + ".").tool_tip = d
        ws.merge_cells(start_row=row_h1, start_column=col_idx, end_row=row_h2, end_column=col_idx)
        col_idx += 1

    puesto_col = col_idx
    ws.cell(row=row_h1, column=puesto_col, value="Puesto")
    ws.merge_cells(start_row=row_h1, start_column=puesto_col, end_row=row_h2, end_column=puesto_col)
    
    for c in range(1, puesto_col + 1):
        for r in [row_h1, row_h2]:
            ws.cell(row=r, column=c).font = bold_font
            ws.cell(row=r, column=c).alignment = center_align
            ws.cell(row=r, column=c).fill = header_fill
            ws.cell(row=r, column=c).border = thin_border

    # --- Cuerpo de la Tabla ---
    current_row = row_h2 + 1
    for i, data in enumerate(sabana_data, 1):
        num_filas_estudiante = len(data['filas_notas']) + 1
        start_merge_row = current_row
        
        # Fila de N°, Nombre y Puesto
        ws.cell(row=start_merge_row, column=1, value=i).alignment = center_align
        ws.cell(row=start_merge_row, column=2, value=f"{data['info'].user.last_name} {data['info'].user.first_name}").alignment = left_align
        ws.cell(row=start_merge_row, column=puesto_col, value=data.get('puesto_final')).alignment = center_align

        # Filas de Periodos
        for fila_periodo in data['filas_notas']:
            ws.cell(row=current_row, column=3, value=fila_periodo['periodo_nombre']).font = bold_font
            col_idx = 4
            for celda in fila_periodo['celdas']:
                nota_str = f"{celda['nota_original'] or '-'}"
                if celda.get('nota_recuperacion') is not None:
                    nota_str += f" ({celda['nota_recuperacion']})"
                ws.cell(row=current_row, column=col_idx, value=nota_str).alignment = center_align
                col_idx += 1
            
            prom_str = f"{fila_periodo['promedio_periodo_original']}"
            if fila_periodo.get('promedio_periodo_recuperado') is not None:
                prom_str += f" ({fila_periodo['promedio_periodo_recuperado']})"
            ws.cell(row=current_row, column=prom_col, value=prom_str).alignment = center_align

            for j, count in enumerate(fila_periodo['resumen_desempeno_periodo_list']):
                ws.cell(row=current_row, column=prom_col + 1 + j, value=count).alignment = center_align
            current_row += 1

        # Fila Acumulada
        fila_ac = data['fila_acumulada']
        ws.cell(row=current_row, column=3, value=fila_ac['periodo_nombre'])
        col_idx = 4
        for celda in fila_ac['celdas']:
            ws.cell(row=current_row, column=col_idx, value=celda['nota']).alignment = center_align
            col_idx += 1
        
        prom_ac_str = f"{fila_ac['promedio_display_original']}"
        if fila_ac.get('promedio_display_recuperado') is not None:
            prom_ac_str += f" ({fila_ac['promedio_display_recuperado']})"
        ws.cell(row=current_row, column=prom_col, value=prom_ac_str).alignment = center_align

        for j, count in enumerate(data['resumen_desempeno_acumulado_list']):
            ws.cell(row=current_row, column=prom_col + 1 + j, value=count).alignment = center_align
        
        for c in range(1, puesto_col + 1):
            for r in range(start_merge_row, current_row + 1):
                ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=current_row, column=c).fill = acumulado_fill
            ws.cell(row=current_row, column=c).font = bold_font
        
        current_row += 1

    # --- Ajuste de Ancho de Columnas ---
    ws.column_dimensions['B'].width = 35
    for col_letter in [get_column_letter(c) for c in range(3, puesto_col + 1)]:
        ws.column_dimensions[col_letter].width = 10
    
    # --- Respuesta HTTP ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Sabana_{curso.nombre}_{periodo.get_nombre_display()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response
