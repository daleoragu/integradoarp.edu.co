# notas/views/export_views.py
import csv
from django.http import HttpResponse, HttpResponseNotFound
from django.db.models import Q
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, get_object_or_404

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

from ..models import Curso, Estudiante, FichaEstudiante, Materia, AreaConocimiento

def es_personal_admin(user):
    """
    Verifica si el usuario es superusuario o pertenece al grupo 'Administradores'.
    """
    return user.is_superuser or user.groups.filter(name='Administradores').exists()

# ==============================================================================
# VISTAS DE EXPORTACIÓN Y PLANTILLA DE ESTUDIANTES
# ==============================================================================

@login_required
@user_passes_test(es_personal_admin)
def descargar_plantilla_estudiantes(request):
    """
    Genera y ofrece para descarga una plantilla de Excel (.xlsx) completa para 
    la importación masiva de estudiantes. Esta función es genérica y no necesita
    filtrar por colegio.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria. Instálela con 'pip install openpyxl'.", status=500)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_migracion_estudiantes.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    headers = [
        'NOMBRES', 'APELLIDOS', 'TIPO_DOCUMENTO (CC, TI, RC, CE, OT)', 'NUMERO_DOCUMENTO', 
        'NOMBRE_CURSO', 'FECHA_NACIMIENTO (YYYY-MM-DD)', 'LUGAR_NACIMIENTO', 'EPS', 
        'GRUPO_SANGUINEO (O+, O-, A+, A-, B+, B-, AB+, AB-)', 'ENFERMEDADES_ALERGIAS',
        'NOMBRE_PADRE', 'CELULAR_PADRE', 'NOMBRE_MADRE', 'CELULAR_MADRE',
        'NOMBRE_ACUDIENTE', 'CELULAR_ACUDIENTE', 'EMAIL_ACUDIENTE', 'ESPERA_EN_PORTERIA (SI/NO)',
        'COLEGIO_ANTERIOR', 'GRADO_ANTERIOR'
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        ws.column_dimensions[get_column_letter(col_num)].width = 22
    
    wb.save(response)
    return response


@login_required
@user_passes_test(es_personal_admin)
def exportar_estudiantes_excel(request):
    """
    Exporta la lista COMPLETA de estudiantes a un archivo Excel, pero únicamente
    los que pertenecen al colegio actual.
    """
    if not request.colegio:
        return HttpResponseNotFound("<h1>Colegio no configurado</h1>")
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria.", status=500)

    # CORRECCIÓN: Filtrar estudiantes por el colegio actual.
    estudiantes_qs = Estudiante.objects.filter(colegio=request.colegio).select_related('user', 'curso').prefetch_related('ficha').order_by('curso__nombre', 'user__last_name')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="exportacion_estudiantes_{request.colegio.slug}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes Exportados"

    headers = [
        'NOMBRES', 'APELLIDOS', 'TIPO_DOCUMENTO', 'NUMERO_DOCUMENTO', 'NOMBRE_CURSO', 
        'FECHA_NACIMIENTO', 'LUGAR_NACIMIENTO', 'EPS', 'GRUPO_SANGUINEO', 
        'ENFERMEDADES_ALERGIAS', 'NOMBRE_PADRE', 'CELULAR_PADRE', 'NOMBRE_MADRE', 
        'CELULAR_MADRE', 'NOMBRE_ACUDIENTE', 'CELULAR_ACUDIENTE', 'EMAIL_ACUDIENTE', 
        'ESPERA_EN_PORTERIA', 'COLEGIO_ANTERIOR', 'GRADO_ANTERIOR'
    ]
    
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 22

    for row_num, estudiante in enumerate(estudiantes_qs, 2):
        ficha = getattr(estudiante, 'ficha', None)

        ws.cell(row=row_num, column=1, value=estudiante.user.first_name)
        ws.cell(row=row_num, column=2, value=estudiante.user.last_name)
        ws.cell(row=row_num, column=5, value=estudiante.curso.nombre if estudiante.curso else '')
        
        if ficha:
            ws.cell(row=row_num, column=3, value=ficha.get_tipo_documento_display())
            ws.cell(row=row_num, column=4, value=ficha.numero_documento)
            ws.cell(row=row_num, column=6, value=ficha.fecha_nacimiento)
            ws.cell(row=row_num, column=7, value=ficha.lugar_nacimiento)
            ws.cell(row=row_num, column=8, value=ficha.eps)
            ws.cell(row=row_num, column=9, value=ficha.get_grupo_sanguineo_display())
            ws.cell(row=row_num, column=10, value=ficha.enfermedades_alergias)
            ws.cell(row=row_num, column=11, value=ficha.nombre_padre)
            ws.cell(row=row_num, column=12, value=ficha.celular_padre)
            ws.cell(row=row_num, column=13, value=ficha.nombre_madre)
            ws.cell(row=row_num, column=14, value=ficha.celular_madre)
            ws.cell(row=row_num, column=15, value=ficha.nombre_acudiente)
            ws.cell(row=row_num, column=16, value=ficha.celular_acudiente)
            ws.cell(row=row_num, column=17, value=ficha.email_acudiente)
            ws.cell(row=row_num, column=18, value="SI" if ficha.espera_en_porteria else "NO")
            ws.cell(row=row_num, column=19, value=ficha.colegio_anterior)
            ws.cell(row=row_num, column=20, value=ficha.grado_anterior)

    wb.save(response)
    return response

# ===============================================================
# VISTAS PARA GESTIÓN DE MATERIAS
# ===============================================================

@login_required
@user_passes_test(es_personal_admin)
def descargar_plantilla_materias(request):
    """
    Genera y ofrece para descarga una plantilla de Excel para la importación de materias.
    Función genérica, no necesita filtrar por colegio.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria.", status=500)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_materias.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Materias"
    headers = ['NOMBRE_MATERIA', 'ABREVIATURA', 'NOMBRE_AREA']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col_num)].width = 30
    
    ws['A2'] = 'EJEMPLO: MATEMÁTICAS'
    ws['B2'] = 'MAT'
    ws['C2'] = 'CIENCIAS EXACTAS'
    
    wb.save(response)
    return response

@login_required
@user_passes_test(es_personal_admin)
def exportar_materias_excel(request):
    """
    Exporta la lista actual de materias y sus áreas a un archivo Excel,
    pero únicamente las del colegio actual.
    """
    if not request.colegio:
        return HttpResponseNotFound("<h1>Colegio no configurado</h1>")
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria.", status=500)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="exportacion_materias_y_areas_{request.colegio.slug}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Materias por Área"
    
    headers = ['NOMBRE_AREA', 'NOMBRE_MATERIA', 'ABREVIATURA']
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 35
    
    row_num = 2
    # CORRECCIÓN: Filtrar áreas y materias por el colegio actual.
    areas = AreaConocimiento.objects.filter(colegio=request.colegio).prefetch_related('materias').order_by('nombre')
    for area in areas:
        # El prefetch ya trae las materias, pero por seguridad podemos re-filtrar en Python.
        materias_del_area_y_colegio = [m for m in area.materias.all() if m.colegio == request.colegio]
        for materia in sorted(materias_del_area_y_colegio, key=lambda m: m.nombre):
            ws.cell(row=row_num, column=1, value=area.nombre)
            ws.cell(row=row_num, column=2, value=materia.nombre)
            ws.cell(row=row_num, column=3, value=materia.abreviatura)
            row_num += 1
        
    wb.save(response)
    return response

# ==============================================================================
# VISTAS PARA PLANTILLA DE DOCENTES
# ==============================================================================

@login_required
@user_passes_test(es_personal_admin)
def descargar_plantilla_docentes(request):
    """
    Genera y ofrece para descarga una plantilla de Excel para la importación de docentes.
    Función genérica, no necesita filtrar por colegio.
    """
    if not EXCEL_SUPPORT:
        return HttpResponse("La librería 'openpyxl' es necesaria para esta función.", status=500)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_docentes.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Docentes"

    headers = ['NOMBRES', 'APELLIDOS', 'DOCUMENTO', 'EMAIL']
    header_font = Font(bold=True)
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    wb.save(response)
    return response
