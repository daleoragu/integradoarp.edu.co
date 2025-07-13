# notas/views/import_views.py
import csv
import io
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import User, Group
from django.db import transaction, IntegrityError
# Se añade HttpResponseNotFound para manejar el caso de un colegio no identificado
from django.http import HttpResponseNotFound
from django.utils.text import slugify
from unidecode import unidecode

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

from ..models.perfiles import Estudiante, Docente, Curso, FichaEstudiante
from ..models.academicos import Materia, AreaConocimiento, PonderacionAreaMateria

def es_superusuario(user):
    return user.is_superuser

@login_required
@user_passes_test(es_superusuario)
def importacion_vista(request):
    """
    Handles the upload and processing of files for bulk data loading,
    ensuring all data is associated with the current school.
    """
    # CORRECCIÓN: Verificar que se ha identificado un colegio.
    if not request.colegio:
        return HttpResponseNotFound("<h1>Colegio no configurado</h1>")

    if request.method == 'POST' and 'archivo_importacion' in request.FILES:
        tipo_importacion = request.POST.get('tipo_importacion')
        archivo = request.FILES['archivo_importacion']

        try:
            with transaction.atomic():
                if tipo_importacion == 'estudiantes':
                    if not EXCEL_SUPPORT:
                        raise Exception("La librería 'openpyxl' es necesaria. Instálela con 'pip install openpyxl'.")
                    if not archivo.name.endswith('.xlsx'):
                        raise Exception("Para importar estudiantes, seleccione un archivo Excel válido (.xlsx).")
                    # CORRECCIÓN: Pasar el objeto colegio a la función de procesamiento.
                    _procesar_excel_estudiantes(request, archivo, request.colegio)
                
                elif tipo_importacion == 'materias':
                    if not EXCEL_SUPPORT:
                        raise Exception("La librería 'openpyxl' es necesaria. Instálela con 'pip install openpyxl'.")
                    if not archivo.name.endswith('.xlsx'):
                         raise Exception(f"Para importar materias, seleccione un archivo Excel válido (.xlsx).")
                    # CORRECCIÓN: Pasar el objeto colegio a la función de procesamiento.
                    _procesar_excel_materias(request, archivo, request.colegio)

                elif tipo_importacion == 'docentes':
                    # CORRECCIÓN: Pasar el objeto colegio a la función de procesamiento.
                    # Asumiendo que se implementará la lógica para docentes en Excel.
                    messages.warning(request, "La importación de docentes aún no está implementada para Excel.")
                    pass
                else:
                    raise Exception("El tipo de importación seleccionado no es válido.")

        except IntegrityError as e:
            messages.error(request, f"Error de integridad: Un dato (como un documento) podría ya existir. Detalles: {e}")
        except Exception as e:
            messages.error(request, f"Error durante el proceso: {e}")

        return redirect(request.META.get('HTTP_REFERER', 'admin_dashboard'))

    return redirect('admin_dashboard')

def _procesar_excel_estudiantes(request, archivo, colegio):
    """Logic to process the student Excel file safely for the current school."""
    creados, errores, omitidos = 0, 0, 0
    grupo_estudiantes, _ = Group.objects.get_or_create(name="Estudiantes")
    
    wb = load_workbook(archivo, data_only=True)
    sheet = wb.active
    map_tipo_doc = {v.upper(): k for k, v in FichaEstudiante.TIPO_DOCUMENTO_CHOICES}
    map_grupo_sang = {v: k for k, v in FichaEstudiante.GRUPO_SANGUINEO_CHOICES}

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        try:
            if not any(row) or not row[0] or not row[1]:
                continue

            row_data = list(row) + [None] * (20 - len(row))
            (nombres, apellidos, tipo_doc_str, num_doc, nombre_curso, fecha_nac_str, 
             lugar_nac, eps, grupo_sang_str, enfermedades, nombre_padre, cel_padre,
             nombre_madre, cel_madre, nombre_acud, cel_acud, email_acud, 
             espera_porteria_str, colegio_ant, grado_ant) = row_data[:20]

            primer_nombre = unidecode(str(nombres).split(' ')[0].lower())
            primer_apellido = unidecode(str(apellidos).split(' ')[0].lower())
            username_base = f"{slugify(primer_nombre)}.{slugify(primer_apellido)}"
            username_final = username_base
            counter = 1
            while User.objects.filter(username=username_final).exists():
                username_final = f"{username_base}{counter}"
                counter += 1
            
            if User.objects.filter(username=username_final).exists():
                omitidos += 1
                continue

            # CORRECCIÓN: Buscar el curso dentro del colegio actual.
            curso = Curso.objects.filter(nombre=str(nombre_curso).strip().upper(), colegio=colegio).first()
            if not curso:
                raise ValueError(f"El curso '{nombre_curso}' no existe en este colegio.")

            user = User.objects.create_user(
                username=username_final, 
                password=username_final,
                first_name=str(nombres or '').strip().upper(), 
                last_name=str(apellidos or '').strip().upper()
            )
            user.groups.add(grupo_estudiantes)
            
            # CORRECCIÓN: Asociar el nuevo estudiante al colegio actual.
            estudiante_obj = Estudiante.objects.create(user=user, curso=curso, colegio=colegio)
            
            fecha_nacimiento = None
            if isinstance(fecha_nac_str, datetime):
                fecha_nacimiento = fecha_nac_str.date()
            elif isinstance(fecha_nac_str, str):
                try:
                    fecha_nacimiento = datetime.strptime(fecha_nac_str, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    fecha_nacimiento = None

            FichaEstudiante.objects.create(
                estudiante=estudiante_obj,
                tipo_documento=map_tipo_doc.get(str(tipo_doc_str).strip().upper(), 'OT') if tipo_doc_str else 'OT',
                numero_documento=str(num_doc).strip() if num_doc else None,
                fecha_nacimiento=fecha_nacimiento,
                lugar_nacimiento=lugar_nac,
                eps=eps,
                grupo_sanguineo=map_grupo_sang.get(str(grupo_sang_str).strip(), None) if grupo_sang_str else None,
                enfermedades_alergias=enfermedades,
                nombre_padre=nombre_padre,
                celular_padre=cel_padre,
                nombre_madre=nombre_madre,
                celular_madre=cel_madre,
                nombre_acudiente=nombre_acud,
                celular_acudiente=cel_acud,
                email_acudiente=email_acud,
                espera_en_porteria=True if espera_porteria_str and 'SI' in str(espera_porteria_str).upper() else False,
                colegio_anterior=colegio_ant,
                grado_anterior=grado_ant
            )
            creados += 1
        except Exception as e:
            messages.warning(request, f"Error en la fila {i} del Excel: {e}")
            errores += 1
            
    messages.success(request, f"Proceso de estudiantes completado. Creados: {creados}. Errores: {errores}. Omitidos (ya existían): {omitidos}.")

def _procesar_excel_materias(request, archivo, colegio):
    """Logic to process the subject Excel file for the current school."""
    creadas, errores, asociaciones_creadas = 0, 0, 0
    wb = load_workbook(archivo, data_only=True)
    sheet = wb.active
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        try:
            if not any(row): continue
            nombre_materia, abreviatura, nombre_area = row[:3]
            
            nombre_materia = str(nombre_materia).strip().upper()
            nombre_area = str(nombre_area).strip().upper()

            if not nombre_materia or not nombre_area:
                continue

            # CORRECCIÓN: Crear o buscar la materia dentro del colegio actual.
            materia_obj, materia_fue_creada = Materia.objects.get_or_create(
                nombre=nombre_materia,
                colegio=colegio,
                defaults={'abreviatura': str(abreviatura).strip().upper() if abreviatura else None}
            )
            if materia_fue_creada:
                creadas += 1

            # CORRECCIÓN: Crear o buscar el área dentro del colegio actual.
            area_obj, _ = AreaConocimiento.objects.get_or_create(nombre=nombre_area, colegio=colegio)

            # CORRECCIÓN: Crear o buscar la ponderación dentro del colegio actual.
            _, ponderacion_fue_creada = PonderacionAreaMateria.objects.get_or_create(
                area=area_obj,
                materia=materia_obj,
                colegio=colegio,
                defaults={'peso_porcentual': 0.00}
            )
            if ponderacion_fue_creada:
                asociaciones_creadas += 1

        except Exception as e:
            messages.warning(request, f"Error en la fila {i} del Excel de materias: {e}")
            errores += 1
            
    messages.success(request, f"Proceso de materias completado. Materias creadas: {creadas}. Asociaciones a áreas creadas: {asociaciones_creadas}. Errores: {errores}.")

def _procesar_csv_docentes(request, reader, colegio):
    """Logic to process the teacher CSV for the current school (preserved)."""
    creados, errores = 0, 0
    grupo_docentes, _ = Group.objects.get_or_create(name="Docentes")
    for i, row in enumerate(reader, 2):
        try:
            if not any(row): continue
            nombres, primer_apellido, segundo_apellido, documento, email, *_ = row
            if not documento: raise ValueError("The document number is mandatory.")
            if User.objects.filter(username=documento).exists(): continue
            user = User.objects.create_user(username=documento, password=documento, email=email, first_name=str(nombres or '').upper(), last_name=f"{str(primer_apellido or '').upper()} {str(segundo_apellido or '').upper()}".strip())
            user.groups.add(grupo_docentes)
            # CORRECCIÓN: Asociar el nuevo docente al colegio actual.
            Docente.objects.create(user=user, colegio=colegio)
            creados += 1
        except Exception as e:
            messages.warning(request, f"Error in teacher CSV row {i}: {e}")
            errores += 1
    messages.success(request, f"Teacher process completed. Created: {creados}. Errors: {errores}.")
