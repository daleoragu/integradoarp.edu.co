# notas/views/importar_asistencia_views.py

from django.shortcuts import redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
import datetime

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Se importan los modelos necesarios para las validaciones
from ..models.academicos import Asistencia, AsignacionDocente
from ..models.perfiles import Estudiante

def es_docente_o_superuser(user):
    """Decorador para asegurar que solo los docentes o superusuarios puedan acceder."""
    return user.is_superuser or user.groups.filter(name='Docentes').exists()

@login_required
@user_passes_test(es_docente_o_superuser)
@transaction.atomic # MEJORA 1: Asegura que toda la importación sea una única transacción.
def importar_asistencia_excel_vista(request):
    """
    Procesa el archivo Excel de plantilla de asistencia subido por un docente.
    """
    if request.method != 'POST':
        messages.error(request, "Método no permitido.")
        return redirect('consulta_asistencia') # Ajusta esta URL si es diferente

    archivo_excel = request.FILES.get('archivo_excel')
    if not archivo_excel:
        messages.error(request, "No se seleccionó ningún archivo.")
        return redirect('consulta_asistencia')

    if not EXCEL_SUPPORT:
        messages.error(request, "El servidor no tiene soporte para leer archivos de Excel (openpyxl no está instalado).")
        return redirect('consulta_asistencia')
    
    # MEJORA 2: Se verifica que el colegio exista en la petición (esencial para multi-colegio)
    if not hasattr(request, 'colegio') or not request.colegio:
        messages.error(request, "Error de configuración: No se pudo identificar el colegio.")
        return redirect('consulta_asistencia')

    try:
        wb = load_workbook(archivo_excel, data_only=True)
        ws = wb.active # Se asume que los datos están en la primera hoja activa

        # Paso 1: Leer los metadatos de la fila 9 para obtener las fechas y la asignación
        metadatos_columnas = {}
        for col in range(3, ws.max_column + 1): # Columnas C, D, E...
            valor_meta = ws.cell(row=9, column=col).value
            if valor_meta and '|' in str(valor_meta):
                try:
                    asignacion_id, fecha_str = str(valor_meta).split('|')
                    # Se valida que la asignación pertenezca al colegio actual
                    if AsignacionDocente.objects.filter(id=int(asignacion_id), colegio=request.colegio).exists():
                        metadatos_columnas[col] = {
                            'asignacion_id': int(asignacion_id),
                            'fecha': datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
                        }
                except (ValueError, TypeError):
                    continue # Ignorar columnas con metadatos malformados
        
        if not metadatos_columnas:
            raise ValueError("El archivo no es una plantilla válida o está corrupto (no se encontraron fechas).")

        registros_creados = 0
        registros_actualizados = 0
        
        # Paso 2: Procesar las filas de estudiantes a partir de la fila 12
        for row in range(12, ws.max_row + 1):
            estudiante_id = ws.cell(row=row, column=2).value # Columna B (oculta)
            if not isinstance(estudiante_id, int):
                continue

            # Se valida que el estudiante pertenezca al colegio actual
            if not Estudiante.objects.filter(id=estudiante_id, colegio=request.colegio).exists():
                continue

            # Paso 3: Iterar sobre las columnas de fechas para registrar las fallas
            for col, meta in metadatos_columnas.items():
                estado_str = ws.cell(row=row, column=col).value
                
                estado_final = None
                justificada = False
                
                if estado_str: # Solo procesamos si la celda no está vacía
                    estado_str = str(estado_str).strip().upper()
                    if estado_str in ('X', 'A'):
                        estado_final = 'A' # Ausente
                    elif estado_str == 'T':
                        estado_final = 'T' # Tarde
                    elif estado_str == 'AJ':
                        estado_final = 'A' # Ausente pero justificada
                        justificada = True
                
                if estado_final:
                    # Se usa update_or_create para manejar registros nuevos y existentes
                    obj, created = Asistencia.objects.update_or_create(
                        colegio=request.colegio, # MEJORA 3: Se asegura que el registro pertenezca al colegio
                        estudiante_id=estudiante_id,
                        asignacion_id=meta['asignacion_id'],
                        fecha=meta['fecha'],
                        defaults={'estado': estado_final, 'justificada': justificada}
                    )
                    if created:
                        registros_creados += 1
                    else:
                        registros_actualizados += 1
        
        messages.success(request, f"Importación completada. Se crearon {registros_creados} y se actualizaron {registros_actualizados} registros de asistencia.")

    except Exception as e:
        messages.error(request, f"Ocurrió un error al procesar el archivo: {e}. Asegúrese de que es la plantilla correcta y no ha sido modificada.")

    return redirect('consulta_asistencia')
